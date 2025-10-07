import pandas as pd
import openpyxl
import time
from llama_cpp import Llama

from utils.processador_narrativa import processar_narrativas, criar_csv_mestre, comparar_com_goldstandard
from utils.similaridade import medir_similaridade
from utils.mapeamento_snomed import prompt_avmap
from utils.processador_excel import carregar_dicionario, salvar_dicionario

# Medir tempo de execução de funções
def medir_tempo(func):
    def wrapper(*args, **kwargs):
        inicio = time.time()
        resultado = func(*args, **kwargs)
        fim = time.time()
        print(f"\nTempo de execução de {func.__name__}: {fim - inicio:.2f} segundos")
        return resultado
    return wrapper

# Pastas e arquivos principais
PASTA_NARRATIVAS = 'narrativas'  # Narrativas XML de entrada
CSV_OUTPUT_FOLDER = 'data/csv_output'  # Saída CSV individual e mestre
DICIONARIO_PATH = 'data/dicionario.json'  # Dicionário SNOMED persistido
RESULTADOS_EXCEL = 'data/Resultados.xlsx'  # Excel final com métricas

# Inicializa lista para armazenar DataFrames de cada narrativa
lista_dataframes_individuais = []

inicio = time.time()  # Marca início da execução total

# Inicializa modelo LLaMA
llm = Llama(
    model_path="modelo/Llama-3.2-3B-Instruct-Q4_K_M.gguf",
    n_ctx=8192,  # Tamanho máximo de contexto
    n_threads=8,  # Ajuste conforme CPU
    n_gpu_layers=20,  # Para acelerar se houver GPU
)

# Processa todas as narrativas XML em blocos, gera CSVs individuais
lista_dataframes_individuais = processar_narrativas(
    PASTA_NARRATIVAS, CSV_OUTPUT_FOLDER, llm, max_tokens=512, temperature=0.0
)

# Cria CSV mestre unindo todos os CSVs individuais
csv_mestre = criar_csv_mestre(lista_dataframes_individuais, CSV_OUTPUT_FOLDER)

if csv_mestre:
    # Compara CSV mestre com gold standard e salva métricas iniciais em Excel
    df_resultado = comparar_com_goldstandard(csv_mestre, PASTA_NARRATIVAS, RESULTADOS_EXCEL)

    # Carrega Excel para análise de similaridade entre termos FP e FN
    workbook = openpyxl.load_workbook(RESULTADOS_EXCEL)
    planilha = workbook['Resultados']

    limiar_similaridade = 0.7  # Limite de similaridade para considerar correspondência. 
    # Isso significa que, se a similaridade calculada entre um termo do prompt e um termo do gold standard for maior que 0.7, eles serão considerados semântica ou textualmente próximos.

    achados_prompt_for_sim = []
    achados_semclin_for_sim = []
    index_semclin = []
    index_prompt = []
    narrativa_anterior = ""

    # Percorre linhas do Excel, realizando análise de similaridade
    for index in range(2, planilha.max_row + 1):
        n1 = ''
        n2 = ''
        if planilha[f'A{index}'].value:
            n1 = str(planilha[f'A{index}'].value)[:4]
        elif planilha[f'G{index}'].value:
            n2 = str(planilha[f'G{index}'].value)[:4]

        narrativa_atual = n1 if n1 else n2

        # Quando muda a narrativa ou chega na última linha, realiza comparação
        if narrativa_atual and (narrativa_atual != narrativa_anterior or index == planilha.max_row):
            if achados_semclin_for_sim and achados_prompt_for_sim:
                for i_p, t_prompt in enumerate(achados_prompt_for_sim):
                    for i_s, t_semclin in enumerate(achados_semclin_for_sim):
                        t_prompt_str = str(t_prompt)
                        t_semclin_str = str(t_semclin)

                        if not t_prompt_str or not t_semclin_str:
                            continue

                        # Calcula similaridade entre termos
                        resultado = medir_similaridade(t_prompt_str, t_semclin_str)

                        if resultado > limiar_similaridade:
                            print(f"\n{resultado:.3f} -> {t_prompt_str} + {t_semclin_str}")

                            # Atualiza classificação para VPP se necessário
                            classificacao_atual = planilha[f'J{index_prompt[i_p]}'].value
                            if classificacao_atual in ['FN', 'FP']:
                                planilha[f'J{index_prompt[i_p]}'] = 'VPP'
                            planilha[f'J{index_semclin[i_s]}'] = ''

                            # Remove termos já processados
                            achados_prompt_for_sim.pop(i_p)
                            achados_semclin_for_sim.pop(i_s)
                            index_prompt.pop(i_p)
                            index_semclin.pop(i_s)
                            break

            # Reseta listas para próxima narrativa
            achados_prompt_for_sim = []
            achados_semclin_for_sim = []
            index_semclin = []
            index_prompt = []

        # Armazena termos FP e FN para análise de similaridade
        if not (index == planilha.max_row and narrativa_atual == narrativa_anterior):
            avaliacao = planilha[f'J{index}'].value
            if avaliacao == 'FN':
                termo = planilha[f'H{index}'].value
                if termo is not None:
                    achados_semclin_for_sim.append(str(termo))
                    index_semclin.append(index)
            elif avaliacao == 'FP':
                termo = planilha[f'D{index}'].value
                if termo is not None:
                    achados_prompt_for_sim.append(str(termo))
                    index_prompt.append(index)

        narrativa_anterior = narrativa_atual

    # Salva Excel atualizado com VPP
    workbook.save(RESULTADOS_EXCEL)

    # Mapeamento SNOMED
    dicionario = carregar_dicionario(DICIONARIO_PATH)
    workbook = openpyxl.load_workbook(RESULTADOS_EXCEL)
    planilha = workbook['Resultados']

    def termo_abreviacao(termo, abreviacao):
        # Junta termo e abreviação para exibição
        return f'{termo} ({abreviacao})'

    # Atualiza coluna de correspondência SNOMED no Excel
    index = 2
    while planilha[f'A{index}'].value or planilha[f'G{index}'].value:
        termo_analisado = planilha[f'D{index}'].value
        sctid_valor = planilha[f'F{index}'].value

        if sctid_valor and sctid_valor != 'NotFound':
            try:
                SCTID = int(sctid_valor)
                abreviacao = planilha[f'E{index}'].value
                termo = termo_abreviacao(termo_analisado, abreviacao) if abreviacao else termo_analisado

                # Atualiza dicionário SNOMED e Excel
                if SCTID in dicionario:
                    if termo in dicionario[SCTID]:
                        planilha[f'K{index}'] = 2
                    else:
                        planilha[f'K{index}'] = 1
                else:
                    resposta = prompt_avmap(SCTID, termo)
                    if resposta == 2:
                        if SCTID in dicionario:
                            dicionario[SCTID].append(termo)
                        else:
                            dicionario[SCTID] = [termo]
                    planilha[f'K{index}'] = resposta
            except ValueError:
                planilha[f'K{index}'] = 'Error'
            except Exception as e:
                planilha[f'K{index}'] = 'Error'

        index += 1

    # Salva Excel e dicionário SNOMED atualizado
    workbook.save(RESULTADOS_EXCEL)
    salvar_dicionario(dicionario, DICIONARIO_PATH)

    # Cálculo das métricas de avaliação
    workbook = openpyxl.load_workbook(RESULTADOS_EXCEL)
    planilha = workbook['Resultados']

    index = 2
    VP = FP = FN = VPP = 0

    # Conta classificações para cálculo de métricas
    while planilha[f'A{index}'].value or planilha[f'G{index}'].value:
        classificacao = planilha[f'J{index}'].value
        if classificacao == 'VP':
            VP += 1
        elif classificacao == 'FP':
            FP += 1
        elif classificacao == 'FN':
            FN += 1
        elif classificacao == 'VPP':
            VPP += 1
        index += 1

    # Calcula precisao, Recall e F1-Score
    precisao = (VP + VPP) / (VP + VPP + FP) if (VP + VPP + FP) > 0 else 0
    recall = (VP + VPP) / (VP + VPP + FN) if (VP + VPP + FN) > 0 else 0
    f1 = 2 * (precisao * recall) / (precisao + recall) if (precisao + recall) > 0 else 0

    metricas = {
        "VP": [VP], "FP": [FP], "FN": [FN], "VPP": [VPP],
        "precisao": [precisao], "Recall": [recall], "F1-Score": [f1]
    }
    df_metricas = pd.DataFrame(metricas)

    # Exibe métricas no terminal
    print("\n+---------------------------------------------------------------+")
    print("|\t 📊 RESULTADOS DA EXTRAÇÃO DE TERMOS CLÍNICOS           |")
    print("+-------+-------+-------+-------+-----------+--------+----------+")
    print("|   VP  |   FP  |   FN  |  VPP  | Precisão  | Recall | F1-Score |")
    print("+-------+-------+-------+-------+-----------+--------+----------+")
    for index, row in df_metricas.iterrows():
        print(f"|{int(row['VP']):5}  |{int(row['FP']):5}  |{int(row['FN']):5}  |{int(row['VPP']):5}  |{row['precisao']:9.3f}  | {row['Recall']:6.3f} |{row['F1-Score']:8.3f}  |")
    print("+-------+-------+-------+-------+-----------+--------+----------+")


    # Contagem SNOMED CT
    index = 2
    resultados = [0, 0, 0]
    while planilha[f'A{index}'].value or planilha[f'G{index}'].value:
        classificacao = planilha[f'K{index}'].value
        if isinstance(classificacao, int):
            resultados[classificacao] += 1
        index += 1

    contagem = {
        "Códigos SNOMED CT não encontrados": [resultados[0]],
        "Códigos existem mas não correspondem": [resultados[1]],
        "Códigos existem e correspondem": [resultados[2]],
        "Total de códigos verificados": [sum(resultados)]
    }

    df = pd.DataFrame(contagem)

    # Exibe resultados do mapeamento SNOMED CT
    print("\n\n+-----------------------------------+-------+")
    print("| 🔍 RESULTADOS DO MAPEAMENTO SNOMED CT     |")
    print("+-----------------------------------+-------+")
    print("| Descrição                         | Total |")
    print("+-----------------------------------+-------+")
    for col in df.columns:
        print(f"| {col:<33} | {int(df[col].iloc[0]):5} |")
    print("+-----------------------------------+-------+")

    fim = time.time()
    tempo_total = fim - inicio
    print(f"\n⏱️   TEMPO TOTAL DE EXECUÇÃO: {tempo_total:.2f} segundos\n")

else:
    print("\nNenhum CSV mestre foi gerado.")
