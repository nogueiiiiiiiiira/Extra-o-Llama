import json
import os
from openpyxl import load_workbook

def verificar_snomed(codigo, termo, dicionario_local):
    # verifica se o código SNOMED corresponde ao termo no dicionário local.

    try:
        codigo_int = int(codigo)
    except ValueError:
        return 0

    if codigo_int in dicionario_local:
        if termo in dicionario_local[codigo_int]:
            return 2  # existe e corresponde
        else:
            return 1  # existe mas não corresponde
    else:
        return 0  # não existe

def termoEAbreviacao(termo, abreviacao):
    return f'{termo} ({abreviacao})' if abreviacao else termo

def mapear_snomed_para_excel(excel_resultados, dicionario_path):
    
    # mapeia SNOMED no Excel e atualiza dicionário.

    try:
        if os.path.exists(dicionario_path):
            with open(dicionario_path, 'r', encoding="utf-8") as f:
                dicionario = json.load(f)
        else:
            dicionario = {}

        workbook = load_workbook(excel_resultados)
        if "Resultados" in workbook.sheetnames:
            planilha = workbook["Resultados"]
        else:
            planilha = workbook.active
            planilha.title = "Resultados"

        index = 2
        while True:
            cell_a = planilha.cell(row=index, column=1).value
            cell_g = planilha.cell(row=index, column=7).value
            if not cell_a and not cell_g:
                break

            termo_analisado = planilha.cell(row=index, column=4).value
            sctid_value = planilha.cell(row=index, column=6).value

            if sctid_value and sctid_value != 'SCTID: NotFound':
                try:
                    SCTID = int(sctid_value)
                    abreviacao = planilha.cell(row=index, column=5).value
                    termo = termoEAbreviacao(termo_analisado, abreviacao) if abreviacao else termo_analisado

                    if SCTID in dicionario:
                        if termo in dicionario[SCTID]:
                            planilha.cell(row=index, column=11).value = 2
                        else:
                            planilha.cell(row=index, column=11).value = 1
                    else:
                        resposta = verificar_snomed(SCTID, termo, dicionario)
                        if resposta == 2:
                            dicionario.setdefault(SCTID, []).append(termo)
                        planilha.cell(row=index, column=11).value = resposta

                except ValueError:
                    print(f"Could not convert SCTID value '{sctid_value}' to int at index {index}. Skipping.")
                    planilha.cell(row=index, column=11).value = 'Error'
                except Exception as e:
                    print(f"An error occurred processing row {index}: {e}")
                    planilha.cell(row=index, column=11).value = 'Error'

            index += 1

        workbook.save(excel_resultados)
        print(f"Excel mapeado e salvo: {excel_resultados}")

        with open(dicionario_path, 'w', encoding="utf-8") as f:
            json.dump(dicionario, f, indent=4, ensure_ascii=False)
        print(f"Dicionário atualizado: {dicionario_path}")

        return dicionario

    except Exception as e:
        print(f"Erro no mapeamento SNOMED: {e}")
        return {}

def contar_resultados_mapeamento(excel_resultados):
    
    # conta os resultados do mapeamento SNOMED.

    try:
        workbook = load_workbook(excel_resultados)
        planilha = workbook['Resultados']

        index = 2
        resultados = [0, 0, 0]
        while planilha[f'A{index}'].value or planilha[f'G{index}'].value:
            classificacao = planilha[f'K{index}'].value
            if classificacao is not None:
                try:
                    resultados[int(classificacao)] += 1
                except (ValueError, IndexError):
                    pass
            index += 1

        contagem = {
            "O código SNOMED CT fornecido não existe": resultados[0],
            "O código existe, mas não corresponde ao termo fornecido": resultados[1],
            "O código existe E corresponde corretamente ao termo fornecido": resultados[2],
            "Total": sum(resultados)
        }

        print("Resultados do mapeamento SNOMED:")
        for key, value in contagem.items():
            print(f"{key}: {value}")

        return contagem

    except Exception as e:
        print(f"Erro ao contar resultados: {e}")
        return {}
