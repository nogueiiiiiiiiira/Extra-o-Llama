import os
import pandas as pd
import xml.etree.ElementTree as ET
from unidecode import unidecode
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from nltk.stem import RSLPStemmer
import nltk

nltk.download('rslp', quiet=True)

def padronizar_string(string):

    # padroniza uma string para comparação, removendo acentos, convertendo para minúsculas e tratando nulos.
    if isinstance(string, str):
        return unidecode(string.lower().strip())
    else:
        return str(string) if string is not None else ""

def relacoes(root):

    # extrai relações de um XML e retorna um dicionário de relações.
    relacaoDicionario = {}
    relations = root.find('RELATIONS')
    if relations is None:
        return relacaoDicionario
    for rel in relations:
        an1 = rel.get('annotation1')
        an2 = rel.get('annotation2')
        tipo = rel.get('reltype')
        if an1 in relacaoDicionario:
            relacaoDicionario[an1].append({'id_relacionado': an2, 'tipo_relacionamento': tipo})
        else:
            relacaoDicionario[an1] = [{'id_relacionado': an2, 'tipo_relacionamento': tipo}]
    return relacaoDicionario

def tagDesejada(tag):

    # verifica se uma tag é desejada para análise.
    if "Diagnostic Procedure" in tag:
        return False
    return (
        "Sign or Symptom" in tag
        or "Disease or Syndrome" in tag
        or "Body Location or Region" in tag
    )

def dadosRelacionados(dicionarioRelacao, id, root, dado):

    # obtém dados relacionados a uma anotação específica em um XML.
    dadoFinal = ""
    verNegado = False

    anotacaoPrincipal = root.find(f".//annotation[@id='{id}']")
    if anotacaoPrincipal is None:
        return "", False

    tagPrincipal = anotacaoPrincipal.get('tag')

    for key, value_list in dicionarioRelacao.items():
        for value in value_list:
            if id == value['id_relacionado']:
                anotRel = root.find(f".//annotation[@id='{key}']")
                if anotRel is not None and "Diagnostic Procedure" in anotRel.get('tag'):
                    return "", False

    if "Diagnostic Procedure" in tagPrincipal:
        return "", False

    if "Negation" in tagPrincipal:
        verNegado = True

    
    for key, value_list in dicionarioRelacao.items():
        for value in value_list:
            if id == value['id_relacionado']:
                anotRel = root.find(f".//annotation[@id='{key}']")
                if anotRel is None:
                    continue
                tagRel = anotRel.get('tag')
                if "Diagnostic Procedure" in tagRel:
                    continue
                if "Negation" in tagRel:
                    verNegado = True
                if tagDesejada(tagRel) or value['tipo_relacionamento'] == 'negation_of':
                    relacao = padronizar_string(anotRel.get('text')) + " "
                    dadoFinal += relacao
                    if value['tipo_relacionamento'] == 'negation_of':
                        verNegado = True

    dadoFinal += dado
    return dadoFinal, verNegado

def processar_narrativa_para_analise(narrativa_atual, df_prompts, excel_resultados, pasta_xml, similaridade=0.7):
    
    # processa uma narrativa para análise, comparação e similaridade.
    registros_resultado = []

    df_narrativa_atual = df_prompts[df_prompts['nomeNarrativa'] == narrativa_atual].copy()
    achados_prompt = df_narrativa_atual[['termo', 'textoPrompt', 'categoria', 'abreviacao', 'SCTID']].to_dict('records')

    narrativa_semclin = os.path.join(pasta_xml, narrativa_atual)

    try:
        tree = ET.parse(narrativa_semclin)
        root = tree.getroot()
        achados_semclin = []
        relacao = relacoes(root)

        tags = root.find('TAGS')
        if tags is None:
            print(f"Nenhuma tag TAGS encontrada no XML {narrativa_semclin}")
            return registros_resultado

        for annotation in tags:
            specific_annotation = annotation.get('tag')
            id = annotation.get('id')
            dado = padronizar_string(annotation.get('text'))
            dadoFinal, negado = dadosRelacionados(relacao, id, root, dado)

            if dadoFinal != "" and ("Sign or Symptom" in specific_annotation or "Disease or Syndrome" in specific_annotation) and not negado and "Diagnostic Procedure" not in specific_annotation:
                categoria = "Sinal ou Sintoma" if "Sign or Symptom" in specific_annotation else "Doença ou Síndrome"
                achados_semclin.append({
                    "narrativa": narrativa_semclin[-8:],
                    "termo": dadoFinal,
                    "categoria": categoria
                })

        achados_semclin.sort(key=lambda item: item['termo'])
        usado_prompt = [False] * len(achados_prompt)
        usado_semclin = [False] * len(achados_semclin)

        # comparação exata. VP (Verdadeiro Positivo)
        for i, achado in enumerate(achados_prompt):
            for j, achado_xml in enumerate(achados_semclin):
                if not usado_prompt[i] and not usado_semclin[j] and achado["termo"] == achado_xml["termo"]:
                    registros_resultado.append({
                        "nomeNarrativa": narrativa_atual,
                        "textoPrompt": achado.get("textoPrompt"),
                        "categoria": achado.get("categoria"),
                        "termoAnalisado": achado.get("termo"),
                        "abreviacao": achado.get("abreviacao"),
                        "SCTID": achado.get("SCTID"),
                        "semClin_nomeNarrativa": achado_xml.get('narrativa'),
                        "semClin_textoAnalisado": achado_xml.get('termo'),
                        "semClin_categoria": achado_xml.get("categoria"),
                        "classificacao": 'VP'
                    })
                    usado_prompt[i] = True
                    usado_semclin[j] = True
                    break

        # comparação por similaridade. VPP (Verdadeiro Positivo Positivo)
        for i, achado in enumerate(achados_prompt):
            if not usado_prompt[i]:
                registros_resultado.append({
                    "nomeNarrativa": narrativa_atual,
                    "textoPrompt": achado.get("textoPrompt"),
                    "categoria": achado.get("categoria"),
                    "termoAnalisado": achado.get("termo"),
                    "abreviacao": achado.get("abreviacao"),
                    "SCTID": achado.get("SCTID"),
                    "semClin_nomeNarrativa": "",
                    "semClin_textoAnalisado": "",
                    "semClin_categoria": "",
                    "classificacao": 'FP'
                })

        # itens do XML não encontrados no prompt. FN
        for j, achado_xml in enumerate(achados_semclin):
            if not usado_semclin[j]:
                registros_resultado.append({
                    "nomeNarrativa": "",
                    "textoPrompt": "",
                    "categoria": "",
                    "termoAnalisado": "",
                    "abreviacao": "",
                    "SCTID": "",
                    "semClin_nomeNarrativa": achado_xml.get('narrativa'),
                    "semClin_textoAnalisado": achado_xml.get('termo'),
                    "semClin_categoria": achado_xml.get("categoria"),
                    "classificacao": 'FN'
                })

    except FileNotFoundError:
        print(f"\nArquivo XML não encontrado para a narrativa {narrativa_atual}")
    except Exception as e:
        print(f"\nErro ao processar narrativa {narrativa_atual}: {e}")

    return registros_resultado

def calcular_similaridade_e_atualizar_excel(excel_resultados, similaridade=0.7):
    
    # calcula similaridade e atualiza o Excel.
    try:
        workbook = load_workbook(excel_resultados)
        planilha = workbook['Resultados']

        stemmer = RSLPStemmer()

        def stem_frase(frase):
            frase_str = str(frase)
            return " ".join(stemmer.stem(w) for w in frase_str.split())

        achados_prompt_for_sim = []
        achados_semclin_for_sim = []
        index_semclin = []
        index_prompt = []
        previous_narrativa = ""

        for index in range(2, planilha.max_row + 1):
            n1 = ''
            n2 = ''
            if planilha[f'A{index}'].value:
                n1 = str(planilha[f'A{index}'].value)[:4]
            elif planilha[f'G{index}'].value:
                n2 = str(planilha[f'G{index}'].value)[:4]

            current_narrativa = n1 if n1 else n2

            if current_narrativa and current_narrativa != previous_narrativa:

                # processar similaridade para narrativa anterior
                if achados_semclin_for_sim and achados_prompt_for_sim:
                    _processar_similaridade_para_narrativa(achados_prompt_for_sim, achados_semclin_for_sim, index_prompt, index_semclin, planilha, similaridade, stemmer)

                achados_prompt_for_sim = []
                achados_semclin_for_sim = []
                index_semclin = []
                index_prompt = []

            if current_narrativa:
                avaliacao = planilha[f'J{index}'].value
                if avaliacao == 'FN':
                    termo = planilha[f'H{index}'].value
                    if termo is not None:
                        achados_semclin_for_sim.append(str(termo))
                        index_semclin.append(index)
                elif avaliacao == 'FP':
                    termo = planilha[f'D{index}'].value
                    if termo is not None:
                        achados_prompt_for_sim.append(str(termo))
                        index_prompt.append(index)

            previous_narrativa = current_narrativa

        # ultima narrativa
        if achados_semclin_for_sim and achados_prompt_for_sim:
            _processar_similaridade_para_narrativa(achados_prompt_for_sim, achados_semclin_for_sim, index_prompt, index_semclin, planilha, similaridade, stemmer)

        workbook.save(excel_resultados)
        print(f"\nSimilaridade calculada e Excel atualizado: {excel_resultados}")

    except Exception as e:
        print(f"\nErro ao calcular similaridade: {e}")

def _processar_similaridade_para_narrativa(achados_prompt_for_sim, achados_semclin_for_sim, index_prompt, index_semclin, planilha, similaridade, stemmer):
    
    #  função auxiliar para processar similaridade de uma narrativa.

    def stem_frase(frase):
        return " ".join(stemmer.stem(str(w)) for w in str(frase).split())

    pares_a_remover_prompt = []
    pares_a_remover_semclin = []

    for i_p, t_prompt in enumerate(achados_prompt_for_sim):
        for i_s, t_semclin in enumerate(achados_semclin_for_sim):
            t_prompt_str = str(t_prompt)
            t_semclin_str = str(t_semclin)

            if not t_prompt_str or not t_semclin_str:
                continue

            doc1 = stem_frase(t_prompt_str)
            doc2 = stem_frase(t_semclin_str)

            if not doc1 or not doc2:
                result_stem = 0.0
            else:
                vectorizer_stem = TfidfVectorizer()
                try:
                    stem_matrix = vectorizer_stem.fit_transform([doc1, doc2])
                    result_stem = cosine_similarity(stem_matrix[0:1], stem_matrix[1:2])[0][0]
                except ValueError:
                    result_stem = 0.0

            if not t_prompt_str or not t_semclin_str:
                result = 0.0
            else:
                vectorizer = TfidfVectorizer()
                try:
                    tfidf_matrix = vectorizer.fit_transform([t_prompt_str, t_semclin_str])
                    result = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0][0]
                except ValueError:
                    result = 0.0

            resultado = max(result_stem, result)

            if resultado > similaridade:
                print(f"{resultado:.3f} -> {doc1} : {t_prompt_str} + {doc2} : {t_semclin_str}")

                current_classification = planilha[f'J{index_prompt[i_p]}'].value
                if current_classification in ['FN', 'FP']:
                    planilha[f'J{index_prompt[i_p]}'] = 'VPP'
                planilha[f'J{index_semclin[i_s]}'] = ''

                pares_a_remover_prompt.append(i_p)
                pares_a_remover_semclin.append(i_s)
                break

    indices_to_remove_prompt = set(pares_a_remover_prompt)
    indices_to_remove_semclin = set(pares_a_remover_semclin)

    achados_prompt_for_sim[:] = [item for i, item in enumerate(achados_prompt_for_sim) if i not in indices_to_remove_prompt]
    index_prompt[:] = [item for i, item in enumerate(index_prompt) if i not in indices_to_remove_prompt]
    achados_semclin_for_sim[:] = [item for i, item in enumerate(achados_semclin_for_sim) if i not in indices_to_remove_semclin]
    index_semclin[:] = [item for i, item in enumerate(index_semclin) if i not in indices_to_remove_semclin]

def calcular_metricas(excel_resultados):
    # calcula métricas VP (Verdadeiro Positivo), FP (Falso Positivo), FN (Falso Negativo), VPP (Verdadeiro Positivo Positivo), precisão, recall, F1.
    
    try:
        workbook = load_workbook(excel_resultados)
        planilha = workbook.active

        VP = FP = FN = VPP = 0

        for index in range(2, planilha.max_row + 1):
            classificacao = planilha[f'J{index}'].value
            classificacao_normalizada = padronizar_string(classificacao)

            if classificacao_normalizada == 'vp':
                VP += 1
            elif classificacao_normalizada == 'fp':
                FP += 1
            elif classificacao_normalizada == 'fn':
                FN += 1
            elif classificacao_normalizada == 'vpp':
                VPP += 1

        precisao = VP / (VP + FP) if (VP + FP) > 0 else 0
        recall = VP / (VP + FN) if (VP + FN) > 0 else 0
        f1_score = (2 * precisao * recall) / (precisao + recall) if (precisao + recall) > 0 else 0

        print(f'Métricas finais - VP: {VP} FP: {FP} FN: {FN} VPP: {VPP}')
        print(f'Precisão: {precisao:.4f} Recall: {recall:.4f} F1-Score: {f1_score:.4f}')

        return VP, FP, FN, VPP, precisao, recall, f1_score

    except Exception as e:
        print(f"Erro ao calcular métricas: {e}")
        return 0, 0, 0, 0, 0, 0, 0
