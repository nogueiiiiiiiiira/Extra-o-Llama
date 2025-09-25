import os
import pandas as pd
import json
import openpyxl
import xml.etree.ElementTree as ET
from llama_cpp import Llama
import unidecode
from openpyxl import Workbook
import numpy as np
import nltk
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from processador_csv import create_dataframe_and_export_csv
from nltk.stem import RSLPStemmer
import time
nltk.download('rslp')

def normalizar_classificacao(valor):
    if not valor:
        return ""
    return unidecode.unidecode(str(valor).strip().upper())

base_dir = os.path.dirname(os.path.abspath(__file__))

modelo_path = os.path.join(base_dir, "modelo", "Llama-3.2-3B-Instruct-Q4_K_M.gguf")
pasta_xml = os.path.join(base_dir, "narrativas")
arquivo_excel = os.path.join(base_dir, "narrativas", "equivalências.xlsx")
output_dir = os.path.join(base_dir, "csv_output")
os.makedirs(output_dir, exist_ok=True)
csv_path = os.path.join(output_dir, "todas_narrativas_extraidas_ordenado.csv")
llm = Llama(model_path=modelo_path, n_ctx=8192)

if not os.path.exists(csv_path):
    pd.DataFrame(columns=["arquivo", "resposta"]).to_csv(csv_path, index=False)
    print("\nCSV inicial criado!")

def dividir_por_tamanho(texto, max_chars=500):
    blocos = []
    palavras = texto.split()
    bloco = ""
    for palavra in palavras:
        if len(bloco) + len(palavra) + 1 <= max_chars:
            bloco += (" " if bloco else "") + palavra
        else:
            blocos.append(bloco.strip())
            bloco = palavra
    if bloco.strip():
        blocos.append(bloco.strip())
    return blocos

def PesquisaClin_Llama(textoClinico):
    prompt = f"""
    **Objetivo:** Identificar achados clínicos e diagnósticos no texto clínico, classificá-los como "Sinal ou Sintoma" ou "Doença ou Síndrome" seguindo as definições e regras abaixo (com foco restrito em achados clínicos e diagnósticos conforme Regra 8). Para cada informação encontrar o código SNOMED CT (SCTID) correspondente em https://termbrowser.nhs.uk/?perspective=full&conceptId1=404684003&edition=uk-edition&release=v20250604&server=https://termbrowser.nhs.uk/sct-browser-api/snomed&langRefset=999001261000000100,999000691000001104, use "NotFound" se não encontrar, e anote o texto original e gere listas em formato de tuplas detalhadas.

    **Definições:**
    *   **Sinal ou Sintoma**: observação do médico ou relato do paciente de uma condição (Ex: icterícia, dor, febre), incluindo achados de exame físico (Ex: `Mucosas úmidas e hipocoradas`) mas exclui achados auscultatórios ou percussórios detalhados (Ver Regra 8).
    *   **Doença ou Síndrome**: Alteração do estado normal de saúde, diagnosticada clinicamente (Ex: HAS, ICC, DM, DAC, pneumonia). Inclui síndromes reconhecidas (Ex: síndrome metabólica). Exclui achados descritivos de exames complementares (Ver Regra 8).

    **Instruções Detalhadas:**

    1.  **Identificação:** Leia o texto clínico e identifique todos os termos que correspondam às definições de "Sinal ou Sintoma" ou "Doença ou Síndrome", **seguindo estritamente TODAS as Regras e Observações Específicas abaixo, especialmente a Regra 8 restritiva sobre achados de exames e procedimentos**. Preste atenção especial a abreviações médicas comuns (ex: HAS, IAM, ICC, DM, DAC).
    2.  **Processamento do Termo:**
        *   Se o termo for uma abreviação conhecida de uma **Doença/Síndrome permitida** (ex: HAS, DM), use a abreviação como 'TermoPrincipal'. Tente encontrar a expansão e registre-a em 'Abreviação'. Se não conhecida, use a abreviação.
        *   Se o termo for um **Sinal/Sintoma permitido** (ex: dispneia, icterícia) ou uma **Doença/Síndrome permitida não abreviada** (ex: pneumonia), use o termo completo como 'TermoPrincipal'. Registre `None` em 'Abreviação'.
    3.  **Classificação:** Para cada termo principal identificado e **permitido pelas regras**, classifique-o como "Sinal ou Sintoma" ou "Doença ou Síndrome".
    4.  **Busca SNOMED CT:** Tente encontrar o SCTID para o termo principal identificado e permitido. Use "NotFound" se não encontrar.
    5.  **Tratamento de Pontuação:** Ignore pontuação externa. Mantenha a interna se fizer parte de um termo composto *permitido* (raro com a nova Regra 8, exceto talvez em nomes de síndromes).
    6.  **Formatação da Saída (Texto Anotado):** Modifique o texto original inserindo a anotação logo após cada termo **identificado e permitido**, usando o formato EXATO: `[Texto analisado: TermoPrincipal | Abreviação: ExpansaoOuAbrevPropriaOuNone | Categoria: Categoria | SCTID: CodigoOuNotFound]`
        *   *Exemplo Doença Abrev.:* `HAS [Texto analisado: HAS | Abreviação: Hipertensão Arterial Sistêmica | Categoria: Doença ou Síndrome | SCTID: 38341003]`
        *   *Exemplo Sinal:* `icterícia [Texto analisado: icterícia | Abreviação: None | Categoria: Sinal ou Sintoma | SCTID: 18165001]`
    7.  **Formatação da Saída (Listas Resumo em Tuplas):** Após o texto anotado, gere as duas listas de resumo apenas com os termos **identificados e permitidos**, no formato exato:
        *   `Sinais ou Sintomas: ([Termo1 | Abrev1 | Cat1 | SCTID1], ...)`
        *   `Doenças ou Síndromes: ([TermoA | AbrevA | CatA | SCTIDA], ...)`
        *   Use `()` se nenhuma entidade permitida for encontrada para uma categoria.
    8.  **Substituição de Quebras de Linha:** No texto anotado final, substitua `\\n` por espaço.

    **Regras e Observações Específicas:**

    *   **Regra 1 (Não Anotar Normalidade Geral):** NÃO anotar estados *gerais* de saúde normal ou achados normais *isolados* (Ex: `BEG`, `CORADA`, `LOTE`, `MV presente e simétrico`). *Exceção:* Achados compostos como na Regra 7, se o componente anormal for um sinal permitido.
    *   **Regra 2 (Não Anotar Medições/Testes):** NÃO anotar medições, valores de laboratório, ou nomes de testes (Ex: `FE 35%`, `PA 145/95`, `ecocardiograma`, `Tomografia`).
    *   **Regra 3 (Não Anotar Verbos de Evolução):** NÃO anotar verbos como `melhorar`, `piorar`. Anote o conceito clínico *permitido* associado.
    *   **Regra 4 (Termos Compostos - Restrita):** Identifique termos compostos **apenas se** representarem um *único* sinal clínico observável permitido (Regra 7/8) ou uma síndrome/doença nomeada. A Regra 8 prevalece sobre achados de exames.
    *   **Regra 5 (Ignorar "Quadro de" / "Histórico de"):** Anote o conceito principal *se* ele for permitido pelas outras regras (Ex: `Quadro de dispneia` -> Anotar `dispneia`; `Hx de fratura` -> Não anotar, pois 'fratura' isolada pode ser considerada achado de exame). *Nota: Modificado pela Regra 8*.
    *   **Regra 6 (Anotar "Sinais de..."):** Continue anotando expressões como `Sinais clínicos de X` ou `Sinais de X` como um único "Sinal ou Sintoma", *mesmo que X seja uma doença*. Isso é considerado um sinal clínico composto. (Ex: `Sinais clínicos de insuficiência cardíaca`).
    *   **Regra 7 (Termos Compostos com Adjetivos/Achados Normais - Restrita):** Anote termos compostos **apenas se** descreverem um sinal clínico diretamente observável e permitido pela Regra 8.
        *   *Exemplo Permitido:* `Mucosas úmidas e hipocoradas` -> Anotar `Mucosas úmidas e hipocoradas` (Sinal/Sintoma - sinal clínico observável).
        *   *Exemplo NÃO Permitido (agora):* `CPP – MV+ BILATERAL, CREPITANTES EM BASE DIREITA` -> Não anotar (Achado auscultatório detalhado, excluído pela Regra 8).
    *   **Regra 8 (Restrição Severa em Achados de Testes/Exames e Procedimentos):** Esta regra **SOBREPÕE** outras regras quando aplicável a achados de exames complementares, achados específicos de exame físico (como ausculta), ou procedimentos.
        *   **NÃO INCLUIR / NÃO ANOTAR:**
            *   **Achados descritos em exames complementares:** Qualquer detalhe interno de relatórios de ecocardiograma, tomografia, ressonância, exames laboratoriais, etc. (Ex: `folhetos espessados`, `refluxo discreto`, `cúspides calcificadas`, `dupla lesão`, `VE hipertrofiado`, `alt de relaxamento`, `AD aumentado`, `lesão nodular`, `infiltrado pulmonar`, `anemia` *se apenas reportada como valor laboratorial*, `hipocalemia`).
            *   **Termos morfológicos ou funcionais isolados de exames:** `espessados`, `hipertrofiado`, `calcificado`, `alterado`, `discreto`, `acentuado`, `reduzido`, `aumentado`, etc., quando derivados de exames.
            *   **Achados específicos de exame físico não diretamente observáveis como sinal:** Achados de ausculta pulmonar ou cardíaca (Ex: `MV+ BILATERAL, CREPITANTES EM BASE DIREITA`, `MVB+, DIFUSAMENTE DIMINUIDO`, `bulhas normofonéticas`, `sopro sistólico`), achados de percussão (Ex: `macicez`), palpação detalhada de órgãos internos se não for um sinal claro como `hepatomegalia`.
            *   **Nomes de Procedimentos ou Cirurgias:** `angioplastia`, `implante de marcapasso`, `cateterismo` (`cat`), `biópsia`.
            *   **Nomes de Testes ou resultados complexos de testes:** `ecg de repouso com bav de 2 grau mobitz 2`, `eletrocardiograma`.
        *   **INCLUIR / ANOTAR APENAS:**
            *   **Sinais clínicos observáveis ou facilmente verificáveis pelo profissional:** `taquicardia`, `icterícia`, `cianose`, `edema` (Ex: `edema em MMII`), `palidez`, `sudorese`, `dispneia` (observada ou referida), `tosse` (observada ou referida), `febre`, `Mucosas úmidas e hipocoradas`.
            *   **Sintomas referidos pelo paciente:** `dor` (Ex: `dor abdominal`, `dor torácica`), `fadiga`, `náusea`, `vômito`, `tontura`, `cefaleia`, `azia`.
            *   **Doenças diagnosticadas (geralmente crônicas ou agudas estabelecidas):** `Diabetes Mellitus` (`DM`), `Hipertensão Arterial Sistêmica` (`HAS`), `Insuficiência Cardíaca Congestiva` (`ICC`), `Doença Pulmonar Obstrutiva Crônica` (`DPOC`), `Doença Arterial Coronariana` (`DAC`), `pneumonia`, `infarto agudo do miocárdio` (`IAM`).
            *   **Síndromes clinicamente reconhecidas:** `Síndrome metabólica`, `Síndrome de Cushing`, `Síndrome de Down`.
        *   **Exemplos de Aplicação da Regra 8:**
            *   *Texto:* `Ecocardio (13/02/15): ... VMi = folhetos espessados, abertura preservada, refluxo discreto, ... VAo = cúspides calcificadas, com dupla lesão; ... VE = hipertrofiado, ..., alt de relaxamento. AD = aumentado.` -> **Não anotar NADA** deste trecho.
            *   *Texto:* `Mucosas úmidas e hipocoradas` -> **Anotar** `Mucosas úmidas e hipocoradas` [Texto analisado: Mucosas úmidas e hipocoradas | Abreviação: None | Categoria: Sinal ou Sintoma | SCTID: NotFound] (Sinal clínico observável).
            *   *Texto:* `CPP – MV+ BILATERAL, CREPITANTES EM BASE DIREITA` -> **Não anotar**. (Achado auscultatório).
            *   *Texto:* `Sinais clínicos de insuficiência cardíaca` -> **Anotar** `Sinais clínicos de insuficiência cardíaca` [Texto analisado: Sinais clínicos de insuficiência cardíaca | Abreviação: None | Categoria: Sinal ou Sintoma | SCTID: NotFound] (Permitido pela Regra 6).
            *   *Texto:* `MVB+, DIFUSAMENTE DIMUIDO;` -> **Não anotar**. (Achado auscultatório).
            *   *Texto:* `Realizou angioplastia. ECG de repouso com bav de 2 grau mobitz 2. Fez cat. Paciente com DM, HAS e DAC.` -> **Anotar apenas** `DM` [Texto analisado: DM | Abreviação: Diabetes Mellitus | Categoria: Doença ou Síndrome | SCTID: 44054006], `HAS` [Texto analisado: HAS | Abreviação: Hipertensão Arterial Sistêmica | Categoria: Doença ou Síndrome | SCTID: 38341003], `DAC` [Texto analisado: DAC | Abreviação: Doença Arterial Coronariana | Categoria: Doença ou Síndrome | SCTID: 53741008]. Ignorar `angioplastia`, `ECG...`, `cat`.
    *   **Regra 9 (NÃO anotar Conceitos Negados ou Ausentes):** Não anotar conceitos clínicos negados, mesmo que pertença a uma das categorias **permitidas** (sinal clínico observável/referido, sintoma, doença diagnosticada, síndrome).
        *   *Exemplos negados que NÃO devem ser anotado:* `sem tosse`, `nega dor`, `afebril`, `assintomático`.
    *   **Regra 10 (NÃO anotar Medicamentos):** Não anotar medicamentos presentes na narrativa. Nem interpretar e inferir um diagnóstico a partir de um medicamento.
        *   *Exemplos medicamentos que NÃO devem ser anotados:* `ancoron`, `svt`.

    **Restrições Importantes:**
    *   NÃO retorne NENHUMA informação adicional, comentários, explicações ou CUIs.
    *   NÃO retorne NENHUMA categoria diferente de sinais/sintomas ou doenças/síndromes **permitidos pelas regras**.
    *   Retorne APENAS o texto anotado seguido pelas listas de resumo em formato de tupla, conforme especificado, contendo **apenas as entidades permitidas**.
    *   Siga rigorosamente os formatos de anotação e das tuplas de resumo.

    **Exemplo de Execução (Refletindo Novas Regras):**

    ***** Texto original de Exemplo:
    Paciente com HAS e ICC diagnosticada. Apresenta dispneia aos esforços e edema em MMII. Nega dor torácica. Afebril. BEG. Exame Pulmonar: MV diminuído em bases. Ecocardiograma mostrou FE=35% e hipertrofia VE. Ex-tabagista. Realizou angioplastia prévia.

    ***** Saída Esperada:
    Paciente com HAS [Texto analisado: HAS | Abreviação: Hipertensão Arterial Sistêmica | Categoria: Doença ou Síndrome | SCTID: 38341003] e ICC [Texto analisado: ICC | Abreviação: Insuficiência Cardíaca Congestiva | Categoria: Doença ou Síndrome | SCTID: 42343007] diagnosticada. Apresenta dispneia [Texto analisado: dispneia | Abreviação: None | Categoria: Sinal ou Sintoma | SCTID: 267036007] aos esforços e edema em MMII [Texto analisado: edema em MMII | Abreviação: None | Categoria: Sinal ou Sintoma | SCTID: 271808008]. Nega dor torácica. Afebril. BEG. Exame Pulmonar: MV diminuído em bases. Ecocardiograma mostrou FE=35% e hipertrofia VE. Ex-tabagista. Realizou angioplastia prévia.

    Sinais ou Sintomas: ([dispneia | None | Sinal ou Sintoma | 267036007], [edema em MMII | None | Sinal ou Sintoma | 271808008]
    Doenças ou Síndromes: ([HAS | Hipertensão Arterial Sistêmica | Doença ou Síndrome | 38341003], [ICC | Insuficiência Cardíaca Congestiva | Doença ou Síndrome | 42343007])

    ---------- FIM DO EXEMPLO ----------

    **Tarefa:** Agora, aplique TODAS essas definições, instruções e regras **restritivas** ao seguinte documento clínico. Retorne APENAS o texto anotado e as listas de resumo em formato de tupla detalhada no formato especificado, contendo somente as entidades permitidas.

    **Documento Clínico:**
    {textoClinico}
    """

    try:
        result = llm(prompt=prompt, max_tokens=500, temperature=0.7)
        return result["choices"][0]["text"].strip()
    except Exception as e:
        return f"Erro na chamada LLaMA: {e}"

def processar_xml(caminho_xml, max_tokens_por_bloco=2000, sleep_sec=0.5):

    try:
        tree = ET.parse(caminho_xml)
        root = tree.getroot()

        texto = root.find(".//TEXT").text.strip()
        if not texto:
            return f"Erro: XML {caminho_xml} sem conteúdo de texto."

        paragrafos = [p.strip() for p in texto.split("\n\n") if p.strip()]
        blocos = []
        bloco_atual = ""
        for p in paragrafos:

            if len(bloco_atual) + len(p) < max_tokens_por_bloco * 4:
                bloco_atual += ("\n\n" + p) if bloco_atual else p
            else:
                blocos.append(bloco_atual)
                bloco_atual = p
        if bloco_atual:
            blocos.append(bloco_atual)

        respostas = []
        for i, bloco in enumerate(blocos, 1):
            print(f"\n   Processando bloco {i}/{len(blocos)} do XML...")
            resp = PesquisaClin_Llama(bloco) 
            respostas.append(resp)
            time.sleep(sleep_sec)

        texto_final = "\n".join(respostas)
        return texto_final

    except ET.ParseError:
        return f"Erro ao parsear XML: {caminho_xml}"
    except Exception as e:
        return f"Erro inesperado ao processar {caminho_xml}: {e}"

PASTA_NARRATIVAS = "narrativas"

arquivos_xml = []
for root, dirs, files in os.walk(PASTA_NARRATIVAS):
    for f in files:
        if f.lower().endswith('.xml'):
            arquivos_xml.append(os.path.join(root, f))


for arquivo_path in arquivos_xml:
    print(f"\nProcessando {arquivo_path}")


CSV_OUTPUT_FOLDER = "csv_output" 
os.makedirs(CSV_OUTPUT_FOLDER, exist_ok=True)

lista_dataframes_individuais = []

if not arquivos_xml:
    print("\n❌ Nenhum arquivo XML válido encontrado na pasta de narrativas.")
else:
    print(f"\n✅ {len(arquivos_xml)} arquivos XML encontrados. Iniciando processamento...")

for caminho_narrativa in arquivos_xml:
    print(f"\n🔄 Iniciando processamento de {caminho_narrativa}")

    max_retries = 3
    retry_delay_seconds = 2

    for attempt in range(max_retries):
        try:
            if not os.path.exists(caminho_narrativa):
                print(f"\nAttempt {attempt + 1}: Arquivo não encontrado: {caminho_narrativa}. Tentando novamente...")
                time.sleep(retry_delay_seconds)
                continue

            resposta = processar_xml(caminho_narrativa, max_tokens_por_bloco=2000, sleep_sec=0.5)
            print(f"\n✅ Processado {caminho_narrativa} (Attempt {attempt + 1}): {resposta[:100]}...")

            nome_arquivo_csv_individual = os.path.join(CSV_OUTPUT_FOLDER, f"output_{os.path.basename(caminho_narrativa)}.csv")
            dataframe_resultante = create_dataframe_and_export_csv(
                input_text=resposta,
                csv_filename=nome_arquivo_csv_individual,
                narrative_name=os.path.basename(caminho_narrativa)
            )

            if dataframe_resultante is None or dataframe_resultante.empty:
                print(f"\n⚠️ Nenhum dado extraído do arquivo {caminho_narrativa}. Criando DataFrame vazio.")
                dataframe_resultante = pd.DataFrame([{
                    "arquivo": os.path.basename(caminho_narrativa),
                    "textoAnalisado": "",
                    "categoria": "",
                    "SCTID": "",
                    "abrev": ""
                }])

            lista_dataframes_individuais.append(dataframe_resultante)
            print(f"\n📊 DataFrame registrado para {caminho_narrativa}")

        except ET.ParseError:
            print(f"\n⚠️ Erro ao parsear XML: {caminho_narrativa}. Pulando arquivo.")
            break
        except Exception as e:
            print(f"\n⚠️ Erro inesperado em {caminho_narrativa}: {e}. Pulando arquivo.")
            break

if lista_dataframes_individuais:
    df_mestre = pd.concat(lista_dataframes_individuais, ignore_index=True)

    df_mestre_sorted = df_mestre.sort_values(
        by=['nomeNarrativa', 'textoAnalisado'],
        ascending=[True, True],
        key=lambda col: col.str.lower() if col.name == 'textoAnalisado' else col,
        ignore_index=True
    )

    master_csv_filename = os.path.join(CSV_OUTPUT_FOLDER, "todas_narrativas_extraidas_ordenado.csv")
    df_mestre_sorted.to_csv(master_csv_filename, index=False, encoding='utf-8', sep=',')
    print(f"\n✅ CSV mestre gerado: {master_csv_filename}")
else:
    print("\n❌ Nenhum DataFrame individual foi gerado.")


# Extração de termos

def padronizar_string(string):
  if isinstance(string, str):
    return unidecode.unidecode(string.lower().strip())
  else:
    return str(string) if string is not None else ""


def relacoes(root):
    relacaoDicionario = {}
    for rel in root.find('RELATIONS'):
        an1 = rel.get('annotation1')
        an2 = rel.get('annotation2')
        tipo = rel.get('reltype')
        if an1 in relacaoDicionario:
            relacaoDicionario[an1].append({'id_relacionado': an2, 'tipo_relacionamento': tipo})
        else:
            relacaoDicionario[an1] = [{'id_relacionado': an2, 'tipo_relacionamento': tipo}]
    return relacaoDicionario

def tagDesejada(tag):
    if "Diagnostic Procedure" in tag:
        return False
    return (
        "Sign or Symptom" in tag
        or "Disease or Syndrome" in tag
        or "Body Location or Region" in tag
    )

def dadosRelacionados(dicionarioRelacao, id, root, dado):
    dadoFinal = ""
    verNegado = False

    anotacaoPrincipal = root.find(f".//annotation[@id='{id}']")
    if anotacaoPrincipal is None:
        return "", False

    tagPrincipal = anotacaoPrincipal.get('tag')

    for key, value_list in dicionarioRelacao.items():
        for value in value_list:
            if id == value['id_relacionado']:
                anotRel = root.find(f".//annotation[@id='{key}']")
                if anotRel is not None and "Diagnostic Procedure" in anotRel.get('tag'):
                    return "", False 

    if "Diagnostic Procedure" in tagPrincipal:
        return "", False

    if "Negation" in tagPrincipal:
        verNegado = True

    for key, value_list in dicionarioRelacao.items():
        for value in value_list:
            if id == value['id_relacionado']:
                anotRel = root.find(f".//annotation[@id='{key}']")
                if anotRel is None:
                    continue
                tagRel = anotRel.get('tag')
                if "Diagnostic Procedure" in tagRel:
                    continue  
                if "Negation" in tagRel:
                    verNegado = True
                if tagDesejada(tagRel) or value['tipo_relacionamento'] == 'negation_of':
                    relacao = padronizar_string(anotRel.get('text')) + " "
                    dadoFinal += relacao
                    if value['tipo_relacionamento'] == 'negation_of':
                        verNegado = True

    dadoFinal += dado
    return dadoFinal, verNegado

similaridade = 0.7 

excel_prompts = os.path.join(CSV_OUTPUT_FOLDER, "todas_narrativas_extraidas_ordenado.csv")
excel_resultados = f"Resultados.xlsx"
df_prompts = pd.read_csv(excel_prompts, encoding='utf-8')


df_resultado = pd.DataFrame(columns=[
    "nomeNarrativa", "textoPrompt", "categoria", "termoAnalisado",
    "abreviacao", "SCTID", "semClin_nomeNarrativa","semClin_textoAnalisado", "semClin_categoria", "classificacao"
])

achados_prompt = []
col_mapping = {
    'nomeNarrativa': 'nomeNarrativa',
    'textoPrompt': 'textoPrompt',
    'categoria': 'categoria',
    'textoAnalisado': 'termo',  
    'abreviacao': 'abreviacao',
    'SCTID': 'SCTID'
}
df_prompts = df_prompts.rename(columns=col_mapping)

narrativas_unicas = df_prompts['nomeNarrativa'].unique()

registros_resultado = []

for narrativa_atual in narrativas_unicas:

    df_narrativa_atual = df_prompts[df_prompts['nomeNarrativa'] == narrativa_atual].copy()
    achados_prompt = df_narrativa_atual[['termo', 'textoPrompt', 'categoria', 'abreviacao', 'SCTID']].to_dict('records')

    narrativa_semclin = caminho_narrativa

    try:
        tree = ET.parse(narrativa_semclin)
        root = tree.getroot()
        achados_semclin = []
        relacao = relacoes(root)

        for annotation in root.find('TAGS'):
            specific_annotation = annotation.get('tag')
            id = annotation.get('id')
            dado = padronizar_string(annotation.get('text'))
            dadoFinal, negado = dadosRelacionados(relacao, id, root, dado)

            if dadoFinal != "" and ("Sign or Symptom" in specific_annotation or "Disease or Syndrome" in specific_annotation) and not negado and "Diagnostic Procedure" not in specific_annotation:
                categoria = "Sinal ou Sintoma" if "Sign or Symptom" in specific_annotation else "Doença ou Síndrome"
                achados_semclin.append({
                    "narrativa": narrativa_semclin[-8:],
                    "termo": dadoFinal,
                    "categoria": categoria
                })

        achados_semclin.sort(key=lambda item: item['termo'])
        usado_prompt = [False] * len(achados_prompt)
        usado_semclin = [False] * len(achados_semclin)

        # Comparando e classificando
        for i, achado in enumerate(achados_prompt):
            for j, achado_xml in enumerate(achados_semclin):
                if not usado_prompt[i] and not usado_semclin[j] and achado["termo"] == achado_xml["termo"]:
                    registros_resultado.append({
                        "nomeNarrativa": narrativa_atual,
                        "textoPrompt": achado.get("textoPrompt"),
                        "categoria": achado.get("categoria"),
                        "termoAnalisado": achado.get("termo"),
                        "abreviacao": achado.get("abreviacao"),
                        "SCTID": achado.get("SCTID"),
                        "semClin_nomeNarrativa": achado_xml.get('narrativa'),
                        "semClin_textoAnalisado": achado_xml.get('termo'),
                        "semClin_categoria": achado_xml.get("categoria"),
                        "classificacao": 'VP'
                    })
                    usado_prompt[i] = True
                    usado_semclin[j] = True
                    break

        # Achados não usados: FP
        for i, achado in enumerate(achados_prompt):
            if not usado_prompt[i]:
                registros_resultado.append({
                    "nomeNarrativa": narrativa_atual,
                    "textoPrompt": achado.get("textoPrompt"),
                    "categoria": achado.get("categoria"),
                    "termoAnalisado": achado.get("termo"),
                    "abreviacao": achado.get("abreviacao"),
                    "SCTID": achado.get("SCTID"),
                    "semClin_nomeNarrativa": "",
                    "semClin_textoAnalisado": "",
                    "semClin_categoria": "",
                    "classificacao": 'FP'
                })

        # Achados não usados: FN
        for j, achado_xml in enumerate(achados_semclin):
            if not usado_semclin[j]:
                registros_resultado.append({
                    "nomeNarrativa": "",
                    "textoPrompt": "",
                    "categoria": "",
                    "termoAnalisado": "",
                    "abreviacao": "",
                    "SCTID": "",
                    "semClin_nomeNarrativa": achado_xml.get('narrativa'),
                    "semClin_textoAnalisado": achado_xml.get('termo'),
                    "semClin_categoria": achado_xml.get("categoria"),
                    "classificacao": 'FN'
                })

    except FileNotFoundError:
        print(f"\nErro: Arquivo XML não encontrado para a narrativa {narrativa_atual}")
        continue
    except Exception as e:
        print(f"\nErro ao processar narrativa {narrativa_atual}: {e}")
        continue

df_resultado = pd.DataFrame(registros_resultado)
df_resultado.to_excel(excel_resultados, index=False, sheet_name='Resultados')

stemmer = RSLPStemmer() 

def stem_frase(frase):
    frase_str = str(frase)
    return " ".join(stemmer.stem(w) for w in frase_str.split())

workbook = openpyxl.load_workbook(excel_resultados)
planilha = workbook['Resultados']

achados_prompt_for_sim = []
achados_semclin_for_sim = []
index_semclin = []
index_prompt = []
previous_narrativa = "" 

for index in range(2, planilha.max_row + 1):
    n1 = ''
    n2 = ''
    if planilha[f'A{index}'].value:
        n1 = str(planilha[f'A{index}'].value)[:4]
    elif planilha[f'G{index}'].value:
        n2 = str(planilha[f'G{index}'].value)[:4]

    current_narrativa = n1 if n1 else n2

    if current_narrativa and (current_narrativa != previous_narrativa or index == planilha.max_row):

        if index == planilha.max_row and current_narrativa == previous_narrativa:

             avaliacao = planilha[f'J{index}'].value
             if avaliacao == 'FN':
                 termo = planilha[f'H{index}'].value
                 if termo is not None:
                     achados_semclin_for_sim.append(str(termo))
                     index_semclin.append(index)
             elif avaliacao == 'FP':
                 termo = planilha[f'D{index}'].value
                 if termo is not None:
                     achados_prompt_for_sim.append(str(termo))
                     index_prompt.append(index)


        if achados_semclin_for_sim and achados_prompt_for_sim:
            pares_a_remover_prompt = []
            pares_a_remover_semclin = []

            for i_p, t_prompt in enumerate(achados_prompt_for_sim):
                for i_s, t_semclin in enumerate(achados_semclin_for_sim):

                    t_prompt_str = str(t_prompt)
                    t_semclin_str = str(t_semclin)

                    if not t_prompt_str or not t_semclin_str:
                         continue 

                    doc1 = stem_frase(t_prompt_str)
                    doc2 = stem_frase(t_semclin_str)

                    if not doc1 or not doc2:
                        result_stem = 0.0
                    else:
                        vectorizer_stem = TfidfVectorizer()
                        try:
                            stem_matrix = vectorizer_stem.fit_transform([doc1, doc2])
                            result_stem = cosine_similarity(stem_matrix[0:1], stem_matrix[1:2])[0][0]
                        except ValueError:
                             result_stem = 0.0 

                    if not t_prompt_str or not t_semclin_str:
                        result = 0.0
                    else:
                        vectorizer = TfidfVectorizer()
                        try:
                            tfidf_matrix = vectorizer.fit_transform([t_prompt_str, t_semclin_str])
                            result = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0][0]
                        except ValueError:
                            result = 0.0

                    resultado = max(result_stem, result)

                    if resultado > similaridade:
                        print(f"\n{resultado:.3f} -> {doc1} : {t_prompt_str} + {doc2} : {t_semclin_str}")

                        current_classification = planilha[f'J{index_prompt[i_p]}'].value
                        if current_classification in ['FN', 'FP']:
                             planilha[f'J{index_prompt[i_p]}'] = 'VPP'
                        planilha[f'J{index_semclin[i_s]}'] = ''

                        pares_a_remover_prompt.append(i_p)
                        pares_a_remover_semclin.append(i_s)
                        break 

            indices_to_remove_prompt = set(pares_a_remover_prompt)
            indices_to_remove_semclin = set(pares_a_remover_semclin)

            achados_prompt_for_sim = [item for i, item in enumerate(achados_prompt_for_sim) if i not in indices_to_remove_prompt]
            index_prompt = [item for i, item in enumerate(index_prompt) if i not in indices_to_remove_prompt]
            achados_semclin_for_sim = [item for i, item in enumerate(achados_semclin_for_sim) if i not in indices_to_remove_semclin]
            index_semclin = [item for i, item in enumerate(index_semclin) if i not in indices_to_remove_semclin]

        if current_narrativa != previous_narrativa or index == planilha.max_row:
             achados_prompt_for_sim = []
             achados_semclin_for_sim = []
             index_semclin = []
             index_prompt = []

    if not (index == planilha.max_row and current_narrativa == previous_narrativa):
        avaliacao = planilha[f'J{index}'].value
        if avaliacao == 'FN':
            termo = planilha[f'H{index}'].value
            if termo is not None:
                achados_semclin_for_sim.append(str(termo))
                index_semclin.append(index)
        elif avaliacao == 'FP':
            termo = planilha[f'D{index}'].value
            if termo is not None:
                achados_prompt_for_sim.append(str(termo))
                index_prompt.append(index)

    previous_narrativa = current_narrativa

if achados_semclin_for_sim and achados_prompt_for_sim:
    pares_a_remover_prompt = []
    pares_a_remover_semclin = []

    for i_p, t_prompt in enumerate(achados_prompt_for_sim):
        for i_s, t_semclin in enumerate(achados_semclin_for_sim):

            t_prompt_str = str(t_prompt)
            t_semclin_str = str(t_semclin)

            if not t_prompt_str or not t_semclin_str:
                 continue 

            doc1 = stem_frase(t_prompt_str)
            doc2 = stem_frase(t_semclin_str)

            if not doc1 or not doc2:
                 result_stem = 0.0
            else:
                vectorizer_stem = TfidfVectorizer()
                try:
                    stem_matrix = vectorizer_stem.fit_transform([doc1, doc2])
                    result_stem = cosine_similarity(stem_matrix[0:1], stem_matrix[1:2])[0][0]
                except ValueError:
                     result_stem = 0.0

            if not t_prompt_str or not t_semclin_str:
                result = 0.0
            else:
                vectorizer = TfidfVectorizer()
                try:
                    tfidf_matrix = vectorizer.fit_transform([t_prompt_str, t_semclin_str])
                    result = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0][0]
                except ValueError:
                    result = 0.0

            resultado = max(result_stem, result)

            if resultado > similaridade:
                print(f"\n{resultado:.3f} -> {doc1} : {t_prompt_str} + {doc2} : {t_semclin_str}")

                current_classification = planilha[f'J{index_prompt[i_p]}'].value
                if current_classification in ['FN', 'FP']:
                     planilha[f'J{index_prompt[i_p]}'] = 'VPP'
                planilha[f'J{index_semclin[i_s]}'] = '' 

                pares_a_remover_prompt.append(i_p)
                pares_a_remover_semclin.append(i_s)
                break 

    indices_to_remove_prompt = set(pares_a_remover_prompt)
    indices_to_remove_semclin = set(pares_a_remover_semclin)

    achados_prompt_for_sim = [item for i, item in enumerate(achados_prompt_for_sim) if i not in indices_to_remove_prompt]
    index_prompt = [item for i, item in enumerate(index_prompt) if i not in indices_to_remove_prompt]
    achados_semclin_for_sim = [item for i, item in enumerate(achados_semclin_for_sim) if i not in indices_to_remove_semclin]
    index_semclin = [item for i, item in enumerate(index_semclin) if i not in indices_to_remove_semclin]

workbook.save(excel_resultados)

# Visualizar resultados
import unicodedata
import pandas as pd
from openpyxl import load_workbook

def normalizar_string(texto):
    if not texto:
        return None
    texto = str(texto).strip().upper() 
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    return texto

arquivo_excel = os.path.join(base_dir, "Resultados.xlsx")
planilha = load_workbook(arquivo_excel)
aba = planilha.active

VP = FP = FN = VPP = 0

for index in range(2, aba.max_row + 1):  
    classificacao = aba[f'J{index}'].value 
    classificacao_normalizada = normalizar_string(classificacao)

    if classificacao_normalizada == 'VP':
        VP += 1
    elif classificacao_normalizada == 'FP':
        FP += 1
    elif classificacao_normalizada == 'FN':
        FN += 1
    elif classificacao_normalizada == 'VPP':
        VPP += 1

precision = VP / (VP + FP) if (VP + FP) > 0 else 0
recall = VP / (VP + FN) if (VP + FN) > 0 else 0
f1_score = (2 * precision * recall) / (precision + recall) if (precision + recall) > 0 else 0

print('\n--- Métricas finais ---')
print(f'VP: {VP}  FP: {FP}  FN: {FN}  VPP: {VPP}')
print(f'Precision: {precision}')
print(f'Recall: {recall}')
print(f'F1-Score: {f1_score}')

# Mapeamento SNOMED

import json
import requests

def verificarSnomedLocal(codigo, termo, dicionario_local):
    try:
        codigo_int = int(codigo)
    except ValueError:
        return 0

    if codigo_int in dicionario_local:
        if termo in dicionario_local[codigo_int]:
            return 2 
        else:
            return 1  
    else:
        return 0  

arquivo_dicionario = os.path.join(base_dir, "dicionario.json")

if os.path.exists(arquivo_dicionario):
    with open(arquivo_dicionario, 'r', encoding="utf-8") as f:
        dicionario = json.load(f)
else:
    dicionario = {}

print(f"\nDicionário carregado: {dicionario}")

import openpyxl
from openpyxl import Workbook
import os

excel_resultados = os.path.join(base_dir, "Resultados.xlsx")

if os.path.exists(excel_resultados):
    workbook = openpyxl.load_workbook(excel_resultados)
    if "Resultados" in workbook.sheetnames:
        planilha = workbook["Resultados"]
    else:
        planilha = workbook.active
        planilha.title = "Resultados"
else:
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Resultados"
    workbook.save(excel_resultados)


def termoEAbreviacao(termo, abreviacao):
  return f'{termo} ({abreviacao})'

index = 2
while True:
    cell_a = planilha.cell(row=index, column=1).value 
    cell_g = planilha.cell(row=index, column=7).value 
    if not cell_a and not cell_g:
        break

    termo_analisado = planilha.cell(row=index, column=4).value  
    sctid_value = planilha.cell(row=index, column=6).value     

    if sctid_value and sctid_value != 'SCTID: NotFound':
        try:
            SCTID = int(sctid_value)
            abreviacao = planilha.cell(row=index, column=5).value  
            termo = termoEAbreviacao(termo_analisado, abreviacao) if abreviacao else termo_analisado

            if SCTID in dicionario:
                if termo in dicionario[SCTID]:
                    planilha.cell(row=index, column=11).value = 2  
                else:
                    planilha.cell(row=index, column=11).value = 1
            else:
                resposta = verificarSnomedLocal(SCTID, termo, dicionario)
                if resposta == 2:
                    dicionario.setdefault(SCTID, []).append(termo)
                planilha.cell(row=index, column=11).value = resposta

        except ValueError:
            print(f"\nCould not convert SCTID value '{sctid_value}' to int at index {index}. Skipping.")
            planilha.cell(row=index, column=11).value = 'Error'
        except Exception as e:
            print(f"\nAn error occurred processing row {index}: {e}")
            planilha.cell(row=index, column=11).value = 'Error'

    index += 1

workbook.save(excel_resultados)
print(f"\nDicionário atualizado: {dicionario}")

with open(arquivo_dicionario, 'w', encoding="utf-8") as f:
    json.dump(dicionario, f, indent=4, ensure_ascii=False)

# Exemplo de execução

termo = 'febre'
codigo = '386661006'

print(f"\n☆ Verificar se '{termo}' corresponde a {codigo} na terminologia SNOMED: ")

print("\n\n☆ Resposta do modelo: ")
resposta = verificarSnomedLocal(codigo, termo, dicionario)
if resposta == '0':
  print("\nO código SNOMED CT fornecido não existe")
elif resposta == '1':
  print("\nO código existe, mas não corresponde ao termo fornecido")
elif resposta == '2':
  print("\nO código existe e corresponde corretamente ao termo fornecido")
else:
  print(resposta)

# Visualizar resultados

excel_resultados = os.path.join(base_dir, "Resultados.xlsx")
workbook = openpyxl.load_workbook(excel_resultados)
planilha = workbook['Resultados']

index = 2

resultados = [0,0,0]
while planilha[f'A{index}'].value or planilha[f'G{index}'].value:
  classificacao = planilha[f'K{index}'].value
  if classificacao is not None:
    if isinstance(classificacao, int):
        resultados[int(classificacao)] += 1
  index += 1

contagem = {
    "O código SNOMED CT fornecido não existe": [resultados[0]],
    "O código existe, mas não corresponde ao termo fornecido": [resultados[1]],
    "O código existe E corresponde corretamente ao termo fornecido": [resultados[2]],
    "Total": [sum(resultados)]
}

df = pd.DataFrame(contagem)

print("\n☆ Resultados do mapeamento SNOMED: ")
print(df)

excel_prompts = os.path.join(base_dir, "csv_output", "todas_narrativas_extraidas_ordenado.csv")
excel_resultados = "Resultados.xlsx"                                
df_prompts = pd.read_csv(excel_prompts,  encoding='utf-8')

achados_prompt = []

col_mapping = {
    'nomeNarrativa': 'nomeNarrativa',
    'textoPrompt': 'textoPrompt',
    'categoria': 'categoria',
    'textoAnalisado': 'termo',  
    'abreviacao': 'abreviacao',
    'SCTID': 'SCTID'
}
df_prompts = df_prompts.rename(columns=col_mapping)

narrativas_unicas = df_prompts['nomeNarrativa'].unique()

for narrativa_atual in narrativas_unicas:

    df_narrativa_atual = df_prompts[df_prompts['nomeNarrativa'] == narrativa_atual].copy()

    achados_prompt = df_narrativa_atual[['termo', 'textoPrompt', 'categoria', 'abreviacao', 'SCTID']].to_dict('records')

    narrativa_semclin = caminho_narrativa

    try:
        tree = ET.parse(narrativa_semclin)
        root = tree.getroot()
        achados_semclin = []
        relacao = relacoes(root)

        for annotation in root.find('TAGS'):
            specific_annotation = annotation.get('tag')
            id = annotation.get('id')
            dado = padronizar_string(annotation.get('text'))
            dadoFinal, negado = dadosRelacionados(relacao, id, root, dado)

            if dadoFinal != "" and ("Sign or Symptom" in specific_annotation or "Disease or Syndrome" in specific_annotation) and not negado and "Diagnostic Procedure" not in specific_annotation:
                categoria = "Sinal ou Sintoma" if "Sign or Symptom" in specific_annotation else "Doença ou Síndrome"
                achados_semclin.append({
                    "narrativa": narrativa_semclin[-8:],
                    "termo": dadoFinal,
                    "categoria": categoria
                })

        achados_semclin.sort(key=lambda item: item['termo'])
        usado_prompt = [False] * len(achados_prompt)
        usado_semclin = [False] * len(achados_semclin)

        for i, achado in enumerate(achados_prompt):
            for j, achado_xml in enumerate(achados_semclin):
                if not usado_prompt[i] and not usado_semclin[j]:
                    if achado["termo"] == achado_xml["termo"]:
                        df_resultado.loc[len(df_resultado)] = {
                            "nomeNarrativa": narrativa_atual,
                            "textoPrompt": achado.get("textoPrompt"),
                            "categoria": achado.get("categoria"),
                            "termoAnalisado": achado.get("termo"),
                            "abreviacao": achado.get("abreviacao"),
                            "SCTID": achado.get("SCTID"),
                            "semClin_nomeNarrativa": achado_xml.get('narrativa'),
                            "semClin_textoAnalisado": achado_xml.get('termo'),
                            "semClin_categoria": achado_xml.get("categoria"),
                            "classificacao": 'VP'
                        }
                        usado_prompt[i] = True
                        usado_semclin[j] = True
                        break

        for i, achado in enumerate(achados_prompt):
            if not usado_prompt[i]:
                df_resultado.loc[len(df_resultado)] = {
                "nomeNarrativa": narrativa_atual,
                "textoPrompt": achado.get("textoPrompt"),
                "categoria": achado.get("categoria"),
                "termoAnalisado": achado.get("termo"),
                "abreviacao": achado.get("abreviacao"),
                "SCTID": achado.get("SCTID"),
                "semClin_nomeNarrativa": "",
                "semClin_textoAnalisado": "",
                "semClin_categoria": "",
                "classificacao": 'FP'
            }

        for j, achado_xml in enumerate(achados_semclin):
            if not usado_semclin[j]:
                df_resultado.loc[len(df_resultado)] = {
                "nomeNarrativa": "",
                "textoPrompt": "",
                "categoria": "",
                "termoAnalisado": "",
                "abreviacao": "",
                "SCTID": "",
                "semClin_nomeNarrativa": achado_xml.get('narrativa'),
                "semClin_textoAnalisado": achado_xml.get('termo'),
                "semClin_categoria": achado_xml.get("categoria"),
                "classificacao": 'FN'
            }

    except FileNotFoundError:
        print(f"\nErro: Arquivo XML não encontrado para a narrativa {narrativa_atual}")
        continue
    except Exception as e:
        print(f"\nErro ao processar narrativa {narrativa_atual}: {e}")
        continue

df_resultado.to_excel(excel_resultados, index=False, sheet_name='Resultados')

stemmer = RSLPStemmer()

def stem_frase(frase):
    return " ".join(stemmer.stem(str(w)) for w in str(frase).split())

workbook = openpyxl.load_workbook(excel_resultados) 
planilha = workbook['Resultados']

achados_prompt_for_sim = []
achados_semclin_for_sim = []
index_semclin = []
index_prompt = []

for index in range(2, planilha.max_row + 1):
    n1 = ''
    n2 = ''
    if planilha[f'A{index}'].value:
        n1 = str(planilha[f'A{index}'].value)[:4]
    elif planilha[f'G{index}'].value:
        n2 = str(planilha[f'G{index}'].value)[:4]

    current_narrativa = n1 if n1 else n2

    if current_narrativa and (current_narrativa != ("" if index == 2 else previous_narrativa)):
        if achados_semclin_for_sim and achados_prompt_for_sim:
            pares_a_remover_prompt = []
            pares_a_remover_semclin = []

            for i_p, t_prompt in enumerate(achados_prompt_for_sim):
                for i_s, t_semclin in enumerate(achados_semclin_for_sim):
                    t_prompt_str = str(t_prompt)
                    t_semclin_str = str(t_semclin)

                    if not t_prompt_str or not t_semclin_str:
                         continue 

                    doc1 = stem_frase(t_prompt_str)
                    doc2 = stem_frase(t_semclin_str)
                    vectorizer_stem = TfidfVectorizer()
                    try:
                        stem_matrix = vectorizer_stem.fit_transform([doc1, doc2])
                        result_stem = cosine_similarity(stem_matrix[0:1], stem_matrix[1:2])[0][0]
                    except ValueError:
                         result_stem = 0.0 

                    vectorizer = TfidfVectorizer()

                    try:
                        tfidf_matrix = vectorizer.fit_transform([t_prompt_str, t_semclin_str])
                        result = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0][0]
                    except ValueError:
                        result = 0.0
                        resultado = max(result_stem, result)

                    if resultado > similaridade:
                        print(f"{resultado:.3f} -> {doc1} : {t_prompt_str} + {doc2} : {t_semclin_str}")

                        planilha[f'J{index_prompt[i_p]}'] = 'VPP'
                        planilha[f'J{index_semclin[i_s]}'] = ''

                        pares_a_remover_prompt.append(i_p)
                        pares_a_remover_semclin.append(i_s)
                        break  

            indices_to_remove_prompt = set(pares_a_remover_prompt)
            indices_to_remove_semclin = set(pares_a_remover_semclin)

            achados_prompt_for_sim = [item for i, item in enumerate(achados_prompt_for_sim) if i not in indices_to_remove_prompt]
            index_prompt = [item for i, item in enumerate(index_prompt) if i not in indices_to_remove_prompt]
            achados_semclin_for_sim = [item for i, item in enumerate(achados_semclin_for_sim) if i not in indices_to_remove_semclin]
            index_semclin = [item for i, item in enumerate(index_semclin) if i not in indices_to_remove_semclin]

        achados_prompt_for_sim = []
        achados_semclin_for_sim = []
        index_semclin = []
        index_prompt = []

    avaliacao = planilha[f'J{index}'].value
    if avaliacao == 'FN':
        termo = planilha[f'H{index}'].value
        if termo is not None:
            achados_semclin_for_sim.append(str(termo))
            index_semclin.append(index)
    elif avaliacao == 'FP':
        termo = planilha[f'D{index}'].value
        if termo is not None:
            achados_prompt_for_sim.append(str(termo))
            index_prompt.append(index)

    previous_narrativa = current_narrativa

if achados_semclin_for_sim and achados_prompt_for_sim:
    pares_a_remover_prompt = []
    pares_a_remover_semclin = []

    for i_p, t_prompt in enumerate(achados_prompt_for_sim):
        for i_s, t_semclin in enumerate(achados_semclin_for_sim):
            t_prompt_str = str(t_prompt)
            t_semclin_str = str(t_semclin)

            if not t_prompt_str or not t_semclin_str:
                 continue 

            doc1 = stem_frase(t_prompt_str)
            doc2 = stem_frase(t_semclin_str)
            vectorizer_stem = TfidfVectorizer()
            try:
                stem_matrix = vectorizer_stem.fit_transform([doc1, doc2])
                result_stem = cosine_similarity(stem_matrix[0:1], stem_matrix[1:2])[0][0]
            except ValueError:
                 result_stem = 0.0

            vectorizer = TfidfVectorizer()
            try:
                tfidf_matrix = vectorizer.fit_transform([t_prompt_str, t_semclin_str])
                result = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0][0]
            except ValueError:
                result = 0.0

            resultado = max(result_stem, result)

            if resultado > similaridade:
                print(f"{resultado:.3f} -> {doc1} : {t_prompt_str} + {doc2} : {t_semclin_str}")

                planilha[f'J{index_prompt[i_p]}'] = 'VPP'
                planilha[f'J{index_semclin[i_s]}'] = ''

                pares_a_remover_prompt.append(i_p)
                pares_a_remover_semclin.append(i_s)
                break  

    indices_to_remove_prompt = set(pares_a_remover_prompt)
    indices_to_remove_semclin = set(pares_a_remover_semclin)

    achados_prompt_for_sim = [item for i, item in enumerate(achados_prompt_for_sim) if i not in indices_to_remove_prompt]
    index_prompt = [item for i, item in enumerate(index_prompt) if i not in indices_to_remove_prompt]
    achados_semclin_for_sim = [item for i, item in enumerate(achados_semclin_for_sim) if i not in indices_to_remove_semclin]
    index_semclin = [item for i, item in enumerate(index_semclin) if i not in indices_to_remove_semclin]

workbook.save(excel_resultados)